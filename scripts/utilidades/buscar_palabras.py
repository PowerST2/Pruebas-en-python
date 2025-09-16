"""
Script para buscar palabras y generar un archivo Excel
Descarga palabras en español y las organiza en un archivo Excel con formato
"""

import os
import requests
import pandas as pd
import time
import openpyxl

def get_words(n=1000):
    """
    Obtiene palabras en español desde un repositorio online
    
    Args:
        n (int): Número de palabras a obtener
        
    Returns:
        list: Lista de palabras filtradas
    """
    try:
        # Usar un diccionario real de palabras en español
        url = 'https://raw.githubusercontent.com/javierarce/palabras/master/listado-general.txt'
        response = requests.get(url)
        words = response.text.splitlines()
        # Filtrar palabras reales, solo letras, sin mayúsculas, sin nombres propios
        words = [w for w in words if w.isalpha() and w.islower() and len(w) > 2]
        words = sorted(set(words))[:n]
        return words
    except Exception as e:
        print(f"Error al obtener palabras: {e}")
        # Palabras de respaldo en caso de error
        return ["python", "programacion", "codigo", "script", "archivo"]

def create_excel(words, filename=None):
    """
    Crea un archivo Excel con las palabras y sus características
    
    Args:
        words (list): Lista de palabras
        filename (str): Nombre del archivo (opcional)
        
    Returns:
        str: Ruta del archivo creado
    """
    if filename is None:
        # Crear archivo en el directorio actual
        filename = 'palabras.xlsx'
    
    data = {
        'N°': list(range(1, len(words)+1)),
        'Palabra': words,
        'Cantidad de letras': [len(w) for w in words]
    }
    
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    return filename

def main():
    """Función principal del script"""
    print("Iniciando proceso de búsqueda de palabras...")
    start = time.time()
    
    # Obtener palabras
    words = get_words(1000)
    print(f"Se obtuvieron {len(words)} palabras")
    
    # Crear archivo Excel
    excel_file = create_excel(words)
    end = time.time()
    duration = int(end - start)
    print(f"Proceso completado en {duration} segundos. Archivo generado: {excel_file}")

    # Abrir el Excel y colorear la celda B240 de amarillo (si existe)
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Verificar que la fila 240 existe
        if len(words) >= 240:
            cell = ws['B240']
            fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.fill = fill
            wb.save(excel_file)
            print('La celda B240 ha sido coloreada de amarillo.')
        else:
            print(f'No se puede colorear B240, solo hay {len(words)} palabras.')
    except Exception as e:
        print(f'Error al colorear celda: {e}')

    # Intentar abrir el archivo Excel automáticamente (solo Windows)
    try:
        if os.name == 'nt':  # Windows
            os.startfile(excel_file)
        else:
            print(f'Archivo generado: {os.path.abspath(excel_file)}')
    except Exception as e:
        print(f'No se pudo abrir el archivo Excel automáticamente: {e}')

if __name__ == "__main__":
    main()