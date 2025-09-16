import os
import requests
import pandas as pd
import time
import openpyxl

def get_words(n=1000):
    # Usar un diccionario real de palabras en español
    url = 'https://raw.githubusercontent.com/javierarce/palabras/master/listado-general.txt'
    response = requests.get(url)
    words = response.text.splitlines()
    # Filtrar palabras reales, solo letras, sin mayúsculas, sin nombres propios
    words = [w for w in words if w.isalpha() and w.islower() and len(w) > 2]
    words = sorted(set(words))[:n]
    return words

def create_excel(words, filename=None):
    if filename is None:
        filename = r'C:\Users\DIAD\Desktop\pruebas\pruebas2\palabras.xlsx'
    data = {
        'N°': list(range(1, len(words)+1)),
        'Palabra': words,
        'Cantidad de letras': [len(w) for w in words]
    }
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    return filename




def main():
    start = time.time()
    words = get_words(1000)
    excel_file = create_excel(words)
    end = time.time()
    duration = int(end - start)
    print(f"Proceso completado en {duration} segundos. Archivo generado: {excel_file}")

    # Abrir el Excel y colorear la celda B240 de amarillo
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    cell = ws['B240']
    fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.fill = fill
    wb.save(excel_file)
    print('La celda B240 ha sido coloreada de amarillo.')

    # Abrir el archivo Excel automáticamente (solo Windows)
    try:
        os.startfile(excel_file)
    except Exception as e:
        print(f'No se pudo abrir el archivo Excel automáticamente: {e}')

if __name__ == "__main__":
    main()
